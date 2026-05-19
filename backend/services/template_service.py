"""
Programmatic Excel template generator.
All file I/O runs inside a ThreadPoolExecutor so it is Windows ProactorEventLoop-safe.
"""

import io
import asyncio
from concurrent.futures import ThreadPoolExecutor
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="0C1F30")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SAMPLE_FILL = PatternFill("solid", fgColor="EDF6F4")
_SAMPLE_FONT = Font(name="Calibri", italic=True, color="4A5568", size=10)
_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, headers: list[str], widths: list[int]) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _style_sample(ws, row_data: list, row_num: int = 2) -> None:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = _SAMPLE_FILL
        cell.font = _SAMPLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _THIN


def _build_counterparties() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Counterparties"

    headers = [
        "Company Name", "Tax ID / VAT Number", "Customer Code",
        "Contact Name", "Status", "Phone1", "Phone2", "Mail1", "Mail2",
    ]
    widths = [28, 22, 18, 22, 12, 18, 18, 32, 32]
    sample = [
        "ACME Corporation", "4500123456", "ACC-001",
        "Jane Smith", "active", "+1-555-010-0001", "+1-555-010-0002",
        "accounting@acme.com", "jane.smith@acme.com",
    ]

    _style_header(ws, headers, widths)
    _style_sample(ws, sample)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_master_balances() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Balances"

    headers = ["Company Name", "Customer Code", "Tax ID / VAT Number", "Balance", "Currency"]
    widths  = [28, 18, 22, 16, 10]
    sample  = ["ACME Corporation", "ACC-001", "4500123456", 150000.00, "USD"]

    _style_header(ws, headers, widths)
    _style_sample(ws, sample)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_statement_of_account() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Account"

    headers = ["Customer Code", "Account Name", "Ref No", "Outstanding", "CCY"]
    widths  = [18, 30, 20, 16, 10]
    sample  = ["ACC-001", "ACME Corporation", "INV-2024-001", 45000.00, "USD"]

    _style_header(ws, headers, widths)
    _style_sample(ws, sample)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_internal_statement() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Statement"

    headers = ["Account Name", "Ref No", "Outstanding", "CCY"]
    widths  = [30, 20, 16, 10]
    sample  = ["ACME Corporation", "INV-2024-001", 45000.00, "USD"]

    _style_header(ws, headers, widths)
    _style_sample(ws, sample)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BUILDERS = {
    "counterparties":        _build_counterparties,
    "master_balances":       _build_master_balances,
    "statement_of_account":  _build_statement_of_account,
    "internal_statement":    _build_internal_statement,
}

_FILENAMES = {
    "counterparties":        "lumina_template_counterparties.xlsx",
    "master_balances":       "lumina_template_master_balances.xlsx",
    "statement_of_account":  "lumina_template_statement_of_account.xlsx",
    "internal_statement":    "lumina_template_internal_statement.xlsx",
}


async def generate_template(template_key: str) -> Tuple[bytes, str]:
    """
    Returns (xlsx_bytes, filename).
    Runs the openpyxl builder in a thread pool to stay
    Windows ProactorEventLoop-safe.
    """
    builder = _BUILDERS.get(template_key)
    if not builder:
        raise ValueError(f"Unknown template key: {template_key!r}")

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        data = await loop.run_in_executor(pool, builder)

    return data, _FILENAMES[template_key]
